# -*- coding: utf-8 -*-
"""
order_export.py — формування Excel-файлу заявки за структурою зразка
"Балабан__17_07_26__270_.xlsx".
"""
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)


def sanitize_filename_part(text):
    text = (text or "").strip()
    text = re.sub(r'[\\/:*?"<>|]', "", text)
    return text


def build_filename(buyer_full_name, order_date, order_number):
    """Прізвище_ДД_ММ_РР_(номер).xlsx — за зразком назви файлу з прикладу."""
    surname = buyer_full_name.split()[0] if buyer_full_name.split() else "Заявка"
    surname = sanitize_filename_part(surname)
    date_part = order_date.strftime("%d_%m_%y")
    return f"{surname}__{date_part}__{order_number}_.xlsx"


def generate_order_excel(header, items, output_path):
    """
    header: dict з полями:
        order_number, order_date (datetime), buyer_name, buyer_address,
        responsible, payment_method, sender_phone, recipient_phone,
        carrier, carrier_branch, recipient_address (текст: область+місто),
        recipient_name
    items: список dict {seq_no, name, code, unit, qty, price, sum,
                         weight_unit, weight_total}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявка"

    def merge_set(rng, value=None, bold=False, align=None):
        ws.merge_cells(rng)
        top_left = rng.split(":")[0]
        cell = ws[top_left]
        if value is not None:
            cell.value = value
        if bold:
            cell.font = BOLD
        if align:
            cell.alignment = Alignment(horizontal=align)
        return cell

    merge_set("A1:D1", f"Заявка  № {header['order_number']}", bold=True)
    ws["E1"] = "Дата"
    date_cell = merge_set("F1:G1", header["order_date"])
    date_cell.number_format = "dd.mm.yyyy"

    merge_set("A2:B2", "Покупець     ")
    merge_set("C2:G2", header.get("buyer_name", ""))

    merge_set("A3:B3", "Адреса покупця")
    merge_set("C3:G3", header.get("buyer_address", ""))

    merge_set("A4:B4", "Телефон відправника")
    merge_set("C4:G4", header.get("sender_phone", ""))

    merge_set("A5:B5", "Відповідальний, ПІБ")
    merge_set("C5:D5", header.get("responsible", "ЧСМ"))
    ws["E5"] = "опл"
    merge_set("F5:G5", header.get("payment_method", ""))

    merge_set("A6:B6", "Телефон одержувача")
    merge_set("C6:G6", header.get("recipient_phone", ""))

    merge_set("A7:B7", "Перевізник")
    carrier_text = header.get("carrier", "")
    if header.get("carrier_branch"):
        carrier_text = f"{carrier_text} {header['carrier_branch']}"
    merge_set("C7:G7", carrier_text)

    merge_set("A8:B8", "Адреса одержувача")
    merge_set("C8:G8", header.get("recipient_address", ""))

    merge_set("A9:B9", "Одержувач, ПІБ")
    merge_set("C9:G9", header.get("recipient_name", ""))

    headers_row = 10
    col_titles = [
        ("A", "№П/п"), ("B", "Найменування"), ("C", None),
        ("D", "Од.вим"), ("E", "К-сть"), ("F", "Ціна "),
        ("G", "Сумма"), ("H", "Вага, кг"), ("I", "Вага, всього"),
    ]
    ws.merge_cells(f"B{headers_row}:C{headers_row}")
    for col, title in col_titles:
        if title is None:
            continue
        cell = ws[f"{col}{headers_row}"]
        cell.value = title
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

    total_sum = 0.0
    total_weight = 0.0
    row = headers_row + 1
    for it in items:
        ws[f"A{row}"] = it["seq_no"]
        ws[f"B{row}"] = it["name"]
        ws[f"C{row}"] = it.get("code", "")
        ws[f"D{row}"] = it.get("unit", "")
        ws[f"E{row}"] = it["qty"]
        ws[f"F{row}"] = it["price"]
        ws[f"G{row}"] = it["sum"]
        ws[f"H{row}"] = it.get("weight_unit", "")
        ws[f"I{row}"] = it.get("weight_total", "")
        for col in "ABCDEFGHI":
            ws[f"{col}{row}"].border = BORDER
        total_sum += it["sum"] or 0
        total_weight += it.get("weight_total") or 0
        row += 1

    ws[f"G{row}"] = round(total_sum, 2)
    ws[f"I{row}"] = round(total_weight, 2)
    ws[f"G{row}"].font = BOLD
    ws[f"I{row}"].font = BOLD

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 7
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 9
    ws.column_dimensions["I"].width = 11

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
