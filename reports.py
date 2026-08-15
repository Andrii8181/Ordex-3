# -*- coding: utf-8 -*-
"""
reports.py — допоміжні функції для вкладки "Звіти та аналітика":
експорт таблиці звіту в Excel, підготовка тексту для копіювання
в буфер обміну (можна вставити прямо в Excel), побудова графіків.
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment


def export_table_to_excel(headers, rows, title, output_path):
    """
    headers: список назв колонок
    rows: список кортежів/списків значень (у тому ж порядку, що headers)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Звіт"

    ws.merge_cells(f"A1:{chr(64 + len(headers))}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    header_row = 3
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(bold=True)

    for r_idx, row in enumerate(rows, start=header_row + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def table_to_clipboard_text(headers, rows):
    """Формує текст у форматі TSV — вставляється прямо в Excel як таблиця."""
    lines = ["\t".join(str(h) for h in headers)]
    for row in rows:
        lines.append("\t".join("" if v is None else str(v) for v in row))
    return "\n".join(lines)
